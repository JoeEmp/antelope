import os
from typing import Iterator
from openpyxl import load_workbook, Workbook


class ExcelReader():

    def __init__(self, filename) -> None:
        self.wb = load_workbook(filename, data_only=True)
        self.table = self.wb.sheetnames[0]

    def get_sheet(self, index=0, name=''):
        if name:
            return self.wb.get_sheet_by_name(name)
        else:
            return self.wb.worksheets[index]

    def get_title(self, index=0, name=''):
        sheet = self.get_sheet(index, name)
        return [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column+1)]

    def get_line(self, index=0, name='', offset=0) -> Iterator[list]:
        """返回excel的每一行

        Args:
            index (int, optional): sheet的索引值. Defaults to 0.
            name (str, optional): sheet的名称, 填入后缩略索引值. Defaults to ''.
            offset (int, optional): 忽略第n行, 默认为0不忽略. Defaults to 0.

        Yields:
            Iterator[list]: 每一行
        """
        sheet = self.get_sheet(index, name)
        for row in range(1+offset, sheet.max_row+1):
            yield [sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column+1)]

    def get_json_line(self, index=0, name='', title_index=1, offset=-1) -> Iterator[dict]:
        """返回excel的每一行,json格式,key为指定标题

        Args:
            index (int, optional): sheet的索引值. Defaults to 0.
            name (str, optional): sheet的名称, 填入后缩略索引值. Defaults to ''.
            title_index (int, optional): 默认第一行为标题. Defaults to 1.
            offset (int, optional): 默认为标题的下一行. Defaults to -1.

        Yields:
            Iterator[dict]: 返回excel的每一行,json格式,key为指定标题
        """
        sheet = self.get_sheet(index, name)
        # 标题
        title = [sheet.cell(row=title_index, column=col).value for col in range(
            1, sheet.max_column+1)]
        if -1 == offset:
            offset = title_index
        for row in range(1+offset, sheet.max_row+1):
            line = [sheet.cell(row=row, column=col).value for col in range(
                1, sheet.max_column+1)]
            yield dict(zip(title, line))

    def print_excel(self, index=0, name='', offset=0):
        for line in self.get_line(index, name, offset):
            print(','.join([str(cell) for cell in line]))


class ExcelWirter():

    def save(self, title: list, rows: list, row_type: str, filename: str, sheet_name=''):
        filename = filename if filename.endswith('.xlsx') else filename+'.xlsx'
        if os.path.exists(filename):
            wb = load_workbook(filename)
            sheet_name = sheet_name or 'Sheet%d' % len(wb.sheetnames)
            if sheet_name not in wb.sheetnames:
                sheet = wb.create_sheet(title=sheet_name)
                sheet.append(title)
                cur_row = 2
            else:
                sheet = wb[sheet_name]
                cur_row = sheet.max_row + 1
        else:
            wb = Workbook()
            sheet = wb.active
            sheet.title = sheet_name
            sheet.append(title)
            cur_row = 2
        if 'list' == row_type:
            for row in rows:
                sheet.append(row)
        elif 'dict' == row_type:
            if len(rows) <= 0:
                pass
            else:
                index_dict = {}
                for key in rows[0].keys():
                    index_dict[key] = title.index(key)+1
                for row in rows:
                    for k, v in row.items():
                        sheet.cell(row=cur_row, column=index_dict[k], value=v)
                    cur_row += 1
        wb.save(filename)

    def save_dict_rows(self, title, rows, filename, sheet_name=''):
        self.save(title, rows, row_type='dict',
                  filename=filename, sheet_name=sheet_name)

    def save_list_rows(self, title, rows, filename, sheet_name=''):
        self.save(title, rows, row_type='list',
                  filename=filename, sheet_name=sheet_name)
